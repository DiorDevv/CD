"""Jadval qatorlarini CSV / JSON / XLSX ga aylantirish (eksport uchun).

Qiymatlar frontend `cellText` bilan bir xil ko'rinishда matnlashtiriladi —
select/multi_select variant nomi, user -> username, datetime -> ISO.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date, datetime
from typing import Any, Iterable

from sqlalchemy import Numeric, Text, cast, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.dynamic import ColumnType, DynamicColumn, DynamicRow, DynamicTable
from app.models.user import User

CSV_BOM = "﻿"


def _sort_clause(sort: str | None, columns: list[DynamicColumn]):
    if not sort:
        return [DynamicRow.position.asc().nullslast(), DynamicRow.created_at.asc()]
    key, _, direction = sort.partition(":")
    col = next((c for c in columns if c.key == key), None)
    if col is None:
        return [DynamicRow.created_at.asc()]
    expr = DynamicRow.data[col.key].astext
    if col.type is ColumnType.number:
        expr = cast(DynamicRow.data[col.key].astext, Numeric)
    ordering = expr.desc() if direction.lower() == "desc" else expr.asc()
    return [ordering.nullslast(), DynamicRow.created_at.asc()]


async def fetch_rows(
    db: AsyncSession,
    table: DynamicTable,
    *,
    q: str | None,
    sort: str | None,
    limit: int,
) -> list[DynamicRow]:
    stmt = select(DynamicRow).where(DynamicRow.table_id == table.id)
    if q:
        stmt = stmt.where(cast(DynamicRow.data, Text).ilike(f"%{q.strip()}%"))
    stmt = stmt.order_by(*_sort_clause(sort, list(table.columns))).limit(limit)
    return list((await db.execute(stmt)).scalars().all())


async def usernames_for(db: AsyncSession, rows: list[DynamicRow], columns: list[DynamicColumn]) -> dict[str, str]:
    user_keys = [c.key for c in columns if c.type is ColumnType.user]
    if not user_keys:
        return {}
    ids: set[uuid.UUID] = set()
    for r in rows:
        for k in user_keys:
            v = r.data.get(k)
            if not v:
                continue
            try:
                ids.add(uuid.UUID(str(v)))
            except (ValueError, TypeError):
                continue
    if not ids:
        return {}
    res = await db.execute(select(User.id, User.username).where(User.id.in_(ids)))
    return {str(uid): name for uid, name in res.all()}


def cell_to_text(col: DynamicColumn, value: Any, usernames: dict[str, str]) -> str:
    if value is None or value == "":
        return ""
    t = col.type
    if t is ColumnType.boolean:
        return "Ha" if value else "Yo'q"
    if t is ColumnType.select:
        opt = next((o for o in (col.config or {}).get("options", []) if o["value"] == value), None)
        return str(opt["label"]) if opt else str(value)
    if t is ColumnType.multi_select:
        arr = value if isinstance(value, (list, tuple)) else []
        out = []
        for v in arr:
            opt = next((o for o in (col.config or {}).get("options", []) if o["value"] == v), None)
            out.append(str(opt["label"]) if opt else str(v))
        return ", ".join(out)
    if t is ColumnType.user:
        return usernames.get(str(value), str(value))
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _typed_value(col: DynamicColumn, value: Any, usernames: dict[str, str]) -> Any:
    """XLSX/JSON uchun — raqam/mantiqiy qiymatni o'z tipida qoldiradi."""
    if value is None or value == "":
        return None
    if col.type is ColumnType.number and isinstance(value, (int, float)):
        return value
    if col.type is ColumnType.boolean:
        return bool(value)
    return cell_to_text(col, value, usernames)


# --- Serializatsiya ---------------------------------------------------------


def to_csv(columns: list[DynamicColumn], rows: Iterable[DynamicRow], usernames: dict[str, str]) -> bytes:
    buf = io.StringIO()
    buf.write(CSV_BOM)
    w = csv.writer(buf, lineterminator="\r\n")
    w.writerow([c.label for c in columns])
    for r in rows:
        w.writerow([cell_to_text(c, r.data.get(c.key), usernames) for c in columns])
    return buf.getvalue().encode("utf-8")


def to_json(columns: list[DynamicColumn], rows: Iterable[DynamicRow], usernames: dict[str, str]) -> bytes:
    out = []
    for r in rows:
        out.append(
            {
                "_id": str(r.id),
                "_created_at": r.created_at.isoformat() if r.created_at else None,
                "_updated_at": r.updated_at.isoformat() if r.updated_at else None,
                **{c.label: _typed_value(c, r.data.get(c.key), usernames) for c in columns},
            }
        )
    return json.dumps(out, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def to_xlsx(columns: list[DynamicColumn], rows: Iterable[DynamicRow], usernames: dict[str, str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Eksport"
    ws.append([c.label for c in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        row_vals = []
        for c in columns:
            v = _typed_value(c, r.data.get(c.key), usernames)
            if isinstance(v, (date, datetime)):
                v = v.isoformat()
            row_vals.append(v)
        ws.append(row_vals)
    for i, c in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(48, len(c.label) + 6))
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


SERIALIZERS = {"csv": to_csv, "json": to_json, "xlsx": to_xlsx}
MEDIA_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "json": "application/json",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
